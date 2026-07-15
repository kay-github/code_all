#!/usr/bin/env python3
"""Build the free-source input dataset for the stock YTD snapshot pipeline."""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import math
import os
import re
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import date, datetime, time as clock_time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterable
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook


SHANGHAI = ZoneInfo("Asia/Shanghai")
DATASET_VERSION = "free-stock-ytd-dataset.v1"
RECOVERY_DATASET_VERSION = "free-stock-ytd-recovery.v1"
RECOVERY_AS_OF_PREFIX = "recover:"
MIN_COVERAGE_RATIO = 0.998
MIN_MASTER_COUNTS = {"SH": 2000, "SZ": 2500, "BSE": 100}
BAOSTOCK_RECONNECT_EVERY = 1000
SSE_URL = "https://query.sse.com.cn/sseQuery/commonQuery.do"
SZSE_URL = "https://www.szse.cn/api/report/ShowReport"
BSE_URL = "https://www.bse.cn/nqxxController/nqxxCnzq.do"
SINA_KLINE_URL = (
    "https://quotes.sina.cn/cn/api/jsonp_v2.php/"
    "var%20_{symbol}=/CN_MarketDataService.getKLineData"
)
SINA_QFQ_URL = "https://finance.sina.com.cn/realstock/company/{symbol}/qfq.js"


class FreeStockSourceError(RuntimeError):
    def __init__(self, code: str, message: str, *, source: str | None = None):
        super().__init__(message)
        self.code = code
        self.source = source


@dataclass(frozen=True)
class MasterRecord:
    code: str
    name: str
    exchange: str
    listing_date: str
    board: str

    @property
    def suffix(self) -> str:
        return "BJ" if self.exchange == "BSE" else self.exchange

    @property
    def symbol(self) -> str:
        return f"{self.code}.{self.suffix}"

    @property
    def provider_code(self) -> str:
        prefix = {"SH": "sh", "SZ": "sz", "BSE": "bj"}[self.exchange]
        return f"{prefix}.{self.code}" if self.exchange != "BSE" else f"{prefix}{self.code}"


def normalize_code(value: Any) -> str:
    if value is None:
        raise FreeStockSourceError("MASTER_CODE_MISSING", "security code is missing")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if not math.isfinite(float(value)):
            raise FreeStockSourceError("MASTER_CODE_INVALID", "security code is invalid")
        value = str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    if not re.fullmatch(r"\d{1,6}", text):
        raise FreeStockSourceError("MASTER_CODE_INVALID", "security code is invalid")
    return text.zfill(6)


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    match = re.match(r"^(\d{4})[-/.]?(\d{2})[-/.]?(\d{2})", text)
    if not match:
        raise FreeStockSourceError("DATE_INVALID", "date is invalid")
    parsed = date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return parsed.isoformat()


def parse_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def request_with_retry(
    session: requests.Session,
    method: str,
    url: str,
    *,
    source: str,
    retries: int = 3,
    timeout: float = 20,
    sleep: Callable[[float], None] = time.sleep,
    **kwargs: Any,
) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            response = session.request(method, url, timeout=timeout, **kwargs)
            if response.status_code == 200:
                return response
            last_error = FreeStockSourceError(
                "HTTP_STATUS_ERROR",
                f"{source} returned HTTP {response.status_code}",
                source=source,
            )
        except requests.RequestException as error:
            last_error = error
        if attempt < retries:
            sleep(min(4.0, 0.5 * (2**attempt)))
    raise FreeStockSourceError(
        "SOURCE_REQUEST_FAILED",
        f"{source} request failed after retries",
        source=source,
    ) from last_error


def default_session(referer: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "Chrome/126 Safari/537.36"
            ),
            "Accept": "application/json,text/plain,*/*",
            "Referer": referer,
        }
    )
    return session


def fetch_sse_master(session: requests.Session, timeout: float = 20) -> list[MasterRecord]:
    records: list[MasterRecord] = []
    for stock_type, board in (("1", "MAIN"), ("8", "STAR")):
        response = request_with_retry(
            session,
            "GET",
            SSE_URL,
            source="sse-master",
            timeout=timeout,
            params={
                "STOCK_TYPE": stock_type,
                "REG_PROVINCE": "",
                "CSRC_CODE": "",
                "STOCK_CODE": "",
                "sqlId": "COMMON_SSE_CP_GPJCTPZ_GPLB_GP_L",
                "COMPANY_STATUS": "2,4,5,7,8",
                "type": "inParams",
                "isPagination": "true",
                "pageHelp.cacheSize": "1",
                "pageHelp.beginPage": "1",
                "pageHelp.pageSize": "10000",
                "pageHelp.pageNo": "1",
                "pageHelp.endPage": "1",
            },
        )
        payload = response.json()
        rows = payload.get("result") if isinstance(payload, dict) else None
        if not isinstance(rows, list):
            raise FreeStockSourceError(
                "SSE_MASTER_INVALID", "SSE master response is invalid", source="sse-master"
            )
        for row in rows:
            code = normalize_code(row.get("A_STOCK_CODE"))
            if code.startswith("689"):
                continue
            records.append(
                MasterRecord(
                    code=code,
                    name=str(row.get("SEC_NAME_CN") or "").strip(),
                    exchange="SH",
                    listing_date=normalize_date(row.get("LIST_DATE")),
                    board=board,
                )
            )
    return records


def _header_index(row: Iterable[Any]) -> dict[str, int]:
    return {str(value).strip(): index for index, value in enumerate(row) if value is not None}


def fetch_szse_master(session: requests.Session, timeout: float = 20) -> list[MasterRecord]:
    response = request_with_retry(
        session,
        "GET",
        SZSE_URL,
        source="szse-master",
        timeout=timeout,
        params={
            "SHOWTYPE": "xlsx",
            "CATALOGID": "1110",
            "TABKEY": "tab1",
        },
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    sheet = workbook.active
    # SZSE exports valid rows with a stale A1:A1 worksheet dimension.
    sheet.reset_dimensions()
    rows = list(sheet.iter_rows(values_only=True))
    code_label = "A\u80a1\u4ee3\u7801"
    name_label = "A\u80a1\u7b80\u79f0"
    listing_label = "A\u80a1\u4e0a\u5e02\u65e5\u671f"
    board_label = "\u677f\u5757"
    header_position = next(
        (index for index, row in enumerate(rows) if code_label in _header_index(row)), None
    )
    if header_position is None:
        raise FreeStockSourceError(
            "SZSE_MASTER_INVALID", "SZSE workbook header is missing", source="szse-master"
        )
    columns = _header_index(rows[header_position])
    required = [code_label, name_label, listing_label, board_label]
    if any(label not in columns for label in required):
        raise FreeStockSourceError(
            "SZSE_MASTER_INVALID", "SZSE workbook columns are incomplete", source="szse-master"
        )
    records: list[MasterRecord] = []
    for row in rows[header_position + 1 :]:
        if not row or row[columns[code_label]] in (None, ""):
            continue
        records.append(
            MasterRecord(
                code=normalize_code(row[columns[code_label]]),
                name=str(row[columns[name_label]] or "").strip(),
                exchange="SZ",
                listing_date=normalize_date(row[columns[listing_label]]),
                board=str(row[columns[board_label]] or "").strip() or "UNKNOWN",
            )
        )
    return records


def parse_bse_payload(text: str) -> dict[str, Any]:
    start = text.find("[")
    end = text.rfind("]")
    if start < 0 or end < start:
        raise FreeStockSourceError(
            "BSE_MASTER_INVALID", "BSE response is invalid", source="bse-master"
        )
    payload = json.loads(text[start : end + 1])
    if not isinstance(payload, list) or not payload or not isinstance(payload[0], dict):
        raise FreeStockSourceError(
            "BSE_MASTER_INVALID", "BSE response is invalid", source="bse-master"
        )
    return payload[0]


def fetch_bse_master(session: requests.Session, timeout: float = 20) -> list[MasterRecord]:
    base_form = {
        "typejb": "T",
        "xxfcbj[]": "2",
        "xxzqdm": "",
        "sortfield": "xxzqdm",
        "sorttype": "asc",
    }
    records: list[MasterRecord] = []
    total_pages: int | None = None
    page = 0
    while total_pages is None or page < total_pages:
        response = request_with_retry(
            session,
            "POST",
            BSE_URL,
            source="bse-master",
            timeout=timeout,
            data={**base_form, "page": str(page)},
        )
        payload = parse_bse_payload(response.text)
        if total_pages is None:
            total_pages = int(payload.get("totalPages") or 0)
            if total_pages <= 0 or total_pages > 100:
                raise FreeStockSourceError(
                    "BSE_MASTER_INVALID", "BSE page count is invalid", source="bse-master"
                )
        content = payload.get("content")
        if not isinstance(content, list):
            raise FreeStockSourceError(
                "BSE_MASTER_INVALID", "BSE page content is invalid", source="bse-master"
            )
        for row in content:
            if isinstance(row, dict):
                code = row.get("xxzqdm")
                name = row.get("xxzqjc")
                listing_date = row.get("fxssrq") or row.get("xxgprq")
            elif isinstance(row, list) and len(row) > 40:
                code = row[38]
                name = row[40]
                listing_date = row[0]
            else:
                raise FreeStockSourceError(
                    "BSE_MASTER_INVALID", "BSE master row is incomplete", source="bse-master"
                )
            records.append(
                MasterRecord(
                    code=normalize_code(code),
                    name=str(name or "").strip(),
                    exchange="BSE",
                    listing_date=normalize_date(listing_date),
                    board="BSE",
                )
            )
        page += 1
    return records


def validate_master(records: list[MasterRecord]) -> dict[str, int]:
    symbols: set[str] = set()
    counts = {"SH": 0, "SZ": 0, "BSE": 0}
    for record in records:
        if record.symbol in symbols:
            raise FreeStockSourceError("MASTER_DUPLICATE", "official master contains a duplicate")
        if not record.name:
            raise FreeStockSourceError("MASTER_NAME_MISSING", "official master name is missing")
        symbols.add(record.symbol)
        counts[record.exchange] += 1
    for exchange, minimum in MIN_MASTER_COUNTS.items():
        if counts[exchange] < minimum:
            raise FreeStockSourceError(
                "MASTER_COUNT_TOO_LOW", f"{exchange} official master count is too low"
            )
    return counts


def collect_baostock_rows(result: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    while result.error_code == "0" and result.next():
        rows.append(result.get_row_data())
    if result.error_code != "0":
        raise FreeStockSourceError(
            "BAOSTOCK_QUERY_FAILED", "Baostock query failed", source="baostock"
        )
    return rows


def login_baostock(
    bs: Any,
    *,
    retries: int = 2,
    sleep: Callable[[float], None] = time.sleep,
) -> None:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            result = bs.login()
            if getattr(result, "error_code", None) == "0":
                return
            last_error = RuntimeError("Baostock login returned an error")
        except Exception as error:
            last_error = error
        if attempt < retries:
            sleep(0.5 * (attempt + 1))
    raise FreeStockSourceError(
        "BAOSTOCK_LOGIN_FAILED", "Baostock login failed", source="baostock"
    ) from last_error


def reconnect_baostock(
    bs: Any,
    *,
    sleep: Callable[[float], None] = time.sleep,
) -> None:
    try:
        bs.logout()
    except Exception:
        pass
    login_baostock(bs, sleep=sleep)


def derive_dates(
    calendar_rows: list[list[str]], now: datetime, required_as_of: str | None = None
) -> tuple[str, str, list[str]]:
    open_dates = sorted(normalize_date(row[0]) for row in calendar_rows if len(row) > 1 and row[1] == "1")
    if not open_dates:
        raise FreeStockSourceError("TRADING_CALENDAR_EMPTY", "trading calendar is empty")
    if required_as_of:
        as_of = normalize_date(required_as_of)
        if as_of not in open_dates:
            raise FreeStockSourceError("AS_OF_NOT_OPEN", "required asOf is not an open date")
    else:
        today = now.date().isoformat()
        cutoff_reached = now.timetz().replace(tzinfo=None) >= clock_time(18, 30)
        candidates = [value for value in open_dates if value < today or (cutoff_reached and value == today)]
        if not candidates:
            raise FreeStockSourceError("AS_OF_UNAVAILABLE", "expected asOf is unavailable")
        as_of = candidates[-1]
    previous_year_end = f"{int(as_of[:4]) - 1}-12-31"
    base_candidates = [value for value in open_dates if value <= previous_year_end]
    if not base_candidates:
        raise FreeStockSourceError("BASE_DATE_UNAVAILABLE", "base date is unavailable")
    return base_candidates[-1], as_of, open_dates


def _record_master_fields(record: MasterRecord, source: str, as_of: str) -> dict[str, Any]:
    return {
        "symbol": record.symbol,
        "code": record.code,
        "name": record.name,
        "exchange": record.exchange,
        "board": record.board,
        "listingDate": record.listing_date,
        "listingStatus": "LISTED",
        "securityType": "A_SHARE",
        "source": source,
        "sourceAsOf": as_of,
    }


def select_endpoints(rows: list[list[str]], base_date: str, as_of: str) -> tuple[tuple[str, float] | None, tuple[str, float] | None]:
    prices: list[tuple[str, float]] = []
    for row in rows:
        if len(row) < 3:
            continue
        value = parse_number(row[2])
        if value is None or value <= 0:
            continue
        prices.append((normalize_date(row[0]), value))
    prices.sort(key=lambda item: item[0])
    base = next((item for item in reversed(prices) if item[0] <= base_date), None)
    current = next((item for item in reversed(prices) if item[0] <= as_of), None)
    return base, current


def baostock_computed_record(
    bs: Any,
    master: MasterRecord,
    base_date: str,
    as_of: str,
    *,
    retries: int = 2,
    sleep: Callable[[float], None] = time.sleep,
) -> dict[str, Any]:
    result = _record_master_fields(master, "baostock", as_of)
    if master.listing_date > base_date:
        result["ineligibilityReason"] = "NEW_LISTING"
        return result
    start = max(
        master.listing_date,
        (date.fromisoformat(base_date) - timedelta(days=62)).isoformat(),
    )
    rows: list[list[str]] | None = None
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            query = bs.query_history_k_data_plus(
                master.provider_code,
                "date,code,close,tradestatus",
                start_date=start,
                end_date=as_of,
                frequency="d",
                adjustflag="2",
            )
            rows = collect_baostock_rows(query)
            break
        except Exception as error:  # provider errors need a clean retry boundary
            last_error = error
            if attempt < retries:
                sleep(0.5 * (attempt + 1))
    if rows is None:
        raise FreeStockSourceError(
            "BAOSTOCK_HISTORY_FAILED", "Baostock history failed", source="baostock"
        ) from last_error
    base, current = select_endpoints(rows, base_date, as_of)
    if base is None:
        fallback = bs.query_history_k_data_plus(
            master.provider_code,
            "date,code,close,tradestatus",
            start_date=master.listing_date,
            end_date=as_of,
            frequency="d",
            adjustflag="2",
        )
        base, current = select_endpoints(collect_baostock_rows(fallback), base_date, as_of)
    if base is None:
        result["ineligibilityReason"] = "MISSING_BASE_PRICE"
        return result
    if current is None:
        result["ineligibilityReason"] = "MISSING_CURRENT_PRICE"
        return result
    computed_ytd = current[1] / base[1] - 1
    result.update(
        {
            "computedYtd": computed_ytd,
            "basePriceDate": base[0],
            "lastPriceDate": current[0],
            "baseAdjustedClose": base[1],
            "lastAdjustedClose": current[1],
            "adjustmentMethod": "qfq",
        }
    )
    return result


def collect_baostock_records(
    bs: Any,
    masters: list[MasterRecord],
    base_date: str,
    as_of: str,
    *,
    reconnect_every: int = BAOSTOCK_RECONNECT_EVERY,
    progress_every: int = 0,
    sleep: Callable[[float], None] = time.sleep,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    records: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []
    for index, master in enumerate(masters, start=1):
        if reconnect_every and index > 1 and (index - 1) % reconnect_every == 0:
            reconnect_baostock(bs, sleep=sleep)
        try:
            record = baostock_computed_record(
                bs, master, base_date, as_of, sleep=sleep
            )
        except Exception:
            # Query retries cannot repair a dead module-level Baostock session.
            reconnect_baostock(bs, sleep=sleep)
            try:
                record = baostock_computed_record(
                    bs, master, base_date, as_of, sleep=sleep
                )
            except Exception as recovery_error:
                failures.append(
                    {
                        "symbol": master.symbol,
                        "source": "baostock",
                        "code": getattr(
                            recovery_error, "code", "BAOSTOCK_HISTORY_FAILED"
                        ),
                    }
                )
                record = {
                    **_record_master_fields(master, "baostock", as_of),
                    "ineligibilityReason": "DATA_QUALITY_REJECTED",
                }
        records.append(record)
        if progress_every and index % progress_every == 0:
            print(
                json.dumps(
                    {"stage": "baostock", "completed": index, "total": len(masters)},
                    ensure_ascii=True,
                ),
                file=sys.stderr,
                flush=True,
            )
    return records, failures


def parse_sina_history(text: str) -> list[tuple[str, float]]:
    start = text.find("=([")
    if start < 0:
        raise FreeStockSourceError(
            "SINA_HISTORY_INVALID", "Sina history response is invalid", source="sina"
        )
    try:
        rows, _ = json.JSONDecoder().raw_decode(text[start + 2 :].lstrip())
    except (json.JSONDecodeError, TypeError) as error:
        raise FreeStockSourceError(
            "SINA_HISTORY_INVALID", "Sina history response is invalid", source="sina"
        ) from error
    if not isinstance(rows, list):
        raise FreeStockSourceError(
            "SINA_HISTORY_INVALID", "Sina history data is invalid", source="sina"
        )
    prices: list[tuple[str, float]] = []
    for row in rows:
        value = parse_number(row.get("close")) if isinstance(row, dict) else None
        if value is None or value <= 0:
            continue
        prices.append((normalize_date(row.get("day")), value))
    return sorted(prices, key=lambda item: item[0])


def parse_sina_factors(text: str) -> list[tuple[str, float]]:
    start = text.find("=")
    if start < 0:
        raise FreeStockSourceError(
            "SINA_FACTOR_INVALID", "Sina factor response is invalid", source="sina"
        )
    try:
        payload, _ = json.JSONDecoder().raw_decode(text[start + 1 :].lstrip())
    except (json.JSONDecodeError, TypeError) as error:
        raise FreeStockSourceError(
            "SINA_FACTOR_INVALID", "Sina factor response is invalid", source="sina"
        ) from error
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, list):
        raise FreeStockSourceError(
            "SINA_FACTOR_INVALID", "Sina factor data is invalid", source="sina"
        )
    factors: list[tuple[str, float]] = []
    for row in data:
        factor = parse_number(row.get("f")) if isinstance(row, dict) else None
        if factor is None or factor <= 0:
            continue
        factors.append((normalize_date(row.get("d")), factor))
    return sorted(factors, key=lambda item: item[0])


def factor_for_date(factors: list[tuple[str, float]], price_date: str) -> tuple[float, str | None]:
    match = next((item for item in reversed(factors) if item[0] <= price_date), None)
    if match:
        return match[1], match[0]
    if not factors:
        return 1.0, None
    raise FreeStockSourceError(
        "SINA_FACTOR_MISSING", "Sina factor does not cover a price date", source="sina"
    )


def sina_computed_record(
    master: MasterRecord,
    base_date: str,
    as_of: str,
    *,
    timeout: float = 20,
) -> dict[str, Any]:
    result = _record_master_fields(master, "sina", as_of)
    if master.listing_date > base_date:
        result["ineligibilityReason"] = "NEW_LISTING"
        return result
    symbol = master.provider_code
    session = default_session("https://finance.sina.com.cn/")
    history_response = request_with_retry(
        session,
        "GET",
        SINA_KLINE_URL.format(symbol=symbol),
        source="sina-history",
        timeout=timeout,
        params={"symbol": symbol, "scale": "240", "ma": "no", "datalen": "1023"},
    )
    factor_response = request_with_retry(
        session,
        "GET",
        SINA_QFQ_URL.format(symbol=symbol),
        source="sina-factor",
        timeout=timeout,
    )
    prices = parse_sina_history(history_response.text)
    factors = parse_sina_factors(factor_response.text)
    base = next((item for item in reversed(prices) if item[0] <= base_date), None)
    current = next((item for item in reversed(prices) if item[0] <= as_of), None)
    if base is None:
        result["ineligibilityReason"] = "MISSING_BASE_PRICE"
        return result
    if current is None:
        result["ineligibilityReason"] = "MISSING_CURRENT_PRICE"
        return result
    base_qfq_factor, base_effective_date = factor_for_date(factors, base[0])
    last_qfq_factor, last_effective_date = factor_for_date(factors, current[0])
    base_adj_factor = 1 / base_qfq_factor
    last_adj_factor = 1 / last_qfq_factor
    computed_ytd = (current[1] * last_adj_factor) / (base[1] * base_adj_factor) - 1
    result.update(
        {
            "computedYtd": computed_ytd,
            "basePriceDate": base[0],
            "lastPriceDate": current[0],
            "baseRawClose": base[1],
            "baseAdjFactor": base_adj_factor,
            "baseAdjFactorDate": base[0],
            "baseFactorEffectiveDate": base_effective_date,
            "lastRawClose": current[1],
            "lastAdjFactor": last_adj_factor,
            "lastAdjFactorDate": current[0],
            "lastFactorEffectiveDate": last_effective_date,
            "adjustmentMethod": "raw-factor",
        }
    )
    return result


def benchmark_rows(
    bs: Any,
    base_date: str,
    as_of: str,
    *,
    retries: int = 2,
    sleep: Callable[[float], None] = time.sleep,
) -> list[dict[str, Any]]:
    by_date: dict[str, float | None] | None = None
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            query = bs.query_history_k_data_plus(
                "sh.000300",
                "date,code,close",
                start_date=base_date,
                end_date=as_of,
                frequency="d",
                adjustflag="3",
            )
            rows = collect_baostock_rows(query)
            candidate_by_date = {
                normalize_date(row[0]): parse_number(row[2])
                for row in rows
                if len(row) >= 3
            }
            if not candidate_by_date.get(base_date) or not candidate_by_date.get(as_of):
                raise FreeStockSourceError(
                    "CSI300_ENDPOINT_MISSING",
                    "Baostock CSI300 endpoint is missing",
                    source="baostock",
                )
            by_date = candidate_by_date
            break
        except Exception as error:
            last_error = error
            if attempt < retries:
                sleep(0.5 * (attempt + 1))
    if by_date is None:
        if isinstance(last_error, FreeStockSourceError) and last_error.code == "CSI300_ENDPOINT_MISSING":
            raise last_error
        raise FreeStockSourceError(
            "BAOSTOCK_QUERY_FAILED", "Baostock CSI300 query failed", source="baostock"
        ) from last_error
    return [
        {"ts_code": "000300.SH", "trade_date": base_date, "close": by_date[base_date]},
        {"ts_code": "000300.SH", "trade_date": as_of, "close": by_date[as_of]},
    ]


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except FileNotFoundError:
            pass
        raise


def build_dataset(options: argparse.Namespace) -> dict[str, Any]:
    started = time.perf_counter()
    now = options.now
    masters = [
        *fetch_sse_master(default_session("https://www.sse.com.cn/"), options.timeout),
        *fetch_szse_master(
            default_session("https://www.szse.cn/market/product/stock/list/index.html"),
            options.timeout,
        ),
        *fetch_bse_master(default_session("https://www.bse.cn/"), options.timeout),
    ]
    master_counts = validate_master(masters)
    if options.limit_per_exchange:
        limited: list[MasterRecord] = []
        for exchange in ("SH", "SZ", "BSE"):
            limited.extend(
                [record for record in masters if record.exchange == exchange][
                    : options.limit_per_exchange
                ]
            )
        masters = limited

    try:
        import baostock as bs
    except ImportError as error:
        raise FreeStockSourceError(
            "BAOSTOCK_NOT_INSTALLED", "baostock is required", source="baostock"
        ) from error

    login_baostock(bs)
    computed_records: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []
    try:
        calendar_start = f"{now.year - 1}-12-01"
        calendar_requested_end = f"{now.year + 1}-12-31"
        calendar_query = bs.query_trade_dates(
            start_date=calendar_start, end_date=calendar_requested_end
        )
        calendar_rows = collect_baostock_rows(calendar_query)
        base_date, as_of, _ = derive_dates(calendar_rows, now, options.as_of)
        index_rows = benchmark_rows(bs, base_date, as_of)

        sh_sz = [record for record in masters if record.exchange in {"SH", "SZ"}]
        baostock_records, baostock_failures = collect_baostock_records(
            bs,
            sh_sz,
            base_date,
            as_of,
            progress_every=options.progress_every,
        )
        computed_records.extend(baostock_records)
        failures.extend(baostock_failures)

        bse = [record for record in masters if record.exchange == "BSE"]
        with concurrent.futures.ThreadPoolExecutor(max_workers=options.bse_workers) as executor:
            future_map = {
                executor.submit(
                    sina_computed_record,
                    master,
                    base_date,
                    as_of,
                    timeout=options.timeout,
                ): master
                for master in bse
            }
            completed = 0
            for future in concurrent.futures.as_completed(future_map):
                master = future_map[future]
                completed += 1
                try:
                    computed_records.append(future.result())
                except Exception as error:
                    failures.append(
                        {
                            "symbol": master.symbol,
                            "source": "sina",
                            "code": getattr(error, "code", "SINA_HISTORY_FAILED"),
                        }
                    )
                    computed_records.append(
                        {
                            **_record_master_fields(master, "sina", as_of),
                            "ineligibilityReason": "DATA_QUALITY_REJECTED",
                        }
                    )
                if options.progress_every and completed % options.progress_every == 0:
                    print(
                        json.dumps(
                            {"stage": "sina", "completed": completed, "total": len(bse)},
                            ensure_ascii=True,
                        ),
                        file=sys.stderr,
                        flush=True,
                    )
    finally:
        bs.logout()

    computed_records.sort(key=lambda record: record["symbol"])
    expected_universe_count = sum(
        1 for master in masters if master.listing_date <= base_date
    )
    eligible_count = sum(
        1 for record in computed_records if record.get("computedYtd") is not None
    )
    coverage_ratio = (
        eligible_count / expected_universe_count if expected_universe_count else 0
    )
    if not options.limit_per_exchange and coverage_ratio < MIN_COVERAGE_RATIO:
        raise FreeStockSourceError(
            "COMPUTED_YTD_COVERAGE_LOW",
            f"computed coverage {coverage_ratio:.6f} is below {MIN_COVERAGE_RATIO}",
        )
    generated_at = datetime.now(tz=SHANGHAI).astimezone(ZoneInfo("UTC")).isoformat().replace(
        "+00:00", "Z"
    )
    return {
        "version": DATASET_VERSION,
        "diagnosticOnly": bool(options.limit_per_exchange),
        "generatedAt": generated_at,
        "baseDate": base_date,
        "asOf": as_of,
        "expectedUniverseCount": expected_universe_count,
        "masterCounts": master_counts,
        "computedRecords": computed_records,
        "indexRows": index_rows,
        "benchmarkSource": "baostock",
        "tradingCalendar": {
            "coveredFrom": normalize_date(calendar_rows[0][0]),
            "coveredThrough": normalize_date(calendar_rows[-1][0]),
            "rows": [
                {"cal_date": normalize_date(row[0]).replace("-", ""), "is_open": int(row[1])}
                for row in calendar_rows
                if len(row) > 1
            ],
        },
        "quality": {
            "eligibleComputed": eligible_count,
            "coverageRatio": coverage_ratio,
            "failureCount": len(failures),
            "failures": failures[:100],
            "elapsedSeconds": round(time.perf_counter() - started, 3),
        },
    }


def parse_now(value: str | None) -> datetime:
    if not value:
        return datetime.now(tz=SHANGHAI)
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=SHANGHAI)
    return parsed.astimezone(SHANGHAI)


def recovery_as_of(value: str | None) -> str | None:
    if not value or not value.startswith(RECOVERY_AS_OF_PREFIX):
        return None
    candidate = value[len(RECOVERY_AS_OF_PREFIX) :]
    try:
        normalized = normalize_date(candidate)
    except Exception as error:
        raise FreeStockSourceError(
            "RECOVERY_AS_OF_INVALID", "recovery asOf is invalid"
        ) from error
    if normalized != candidate:
        raise FreeStockSourceError(
            "RECOVERY_AS_OF_INVALID", "recovery asOf is invalid"
        )
    return normalized


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True)
    parser.add_argument("--as-of")
    parser.add_argument("--now")
    parser.add_argument("--timeout", type=float, default=20)
    parser.add_argument("--bse-workers", type=int, default=4)
    parser.add_argument("--progress-every", type=int, default=100)
    parser.add_argument("--limit-per-exchange", type=int)
    args = parser.parse_args(argv)
    args.now = parse_now(args.now)
    if args.bse_workers < 1 or args.bse_workers > 8:
        parser.error("--bse-workers must be between 1 and 8")
    if args.limit_per_exchange is not None and args.limit_per_exchange < 1:
        parser.error("--limit-per-exchange must be positive")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_arguments(argv)
    try:
        recover_date = recovery_as_of(args.as_of)
        if recover_date:
            dataset = {
                "version": RECOVERY_DATASET_VERSION,
                "recoveryOnly": True,
                "recoverAsOf": recover_date,
            }
            _atomic_write_json(Path(args.output).resolve(), dataset)
            print(
                json.dumps(
                    {
                        "ok": True,
                        "version": RECOVERY_DATASET_VERSION,
                        "recoveryOnly": True,
                        "asOf": recover_date,
                    },
                    ensure_ascii=True,
                )
            )
            return 0
        dataset = build_dataset(args)
        _atomic_write_json(Path(args.output).resolve(), dataset)
        summary = {
            "ok": True,
            "version": dataset["version"],
            "diagnosticOnly": dataset["diagnosticOnly"],
            "asOf": dataset["asOf"],
            "baseDate": dataset["baseDate"],
            "masterCounts": dataset["masterCounts"],
            "expectedUniverseCount": dataset["expectedUniverseCount"],
            "eligibleComputed": dataset["quality"]["eligibleComputed"],
            "coverageRatio": dataset["quality"]["coverageRatio"],
            "failureCount": dataset["quality"]["failureCount"],
            "elapsedSeconds": dataset["quality"]["elapsedSeconds"],
        }
        print(json.dumps(summary, ensure_ascii=True))
        return 0
    except Exception as error:
        print(
            json.dumps(
                {
                    "ok": False,
                    "error": getattr(error, "code", "FREE_STOCK_WORKER_FAILED"),
                    "source": getattr(error, "source", None),
                },
                ensure_ascii=True,
            ),
            file=sys.stderr,
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
